#
#  download_special_unit_menu.py
#
#   Uso:  <unidade> onde: <unidade> = _id mongodb da unidade
#

from datetime import datetime
import os
import pymongo
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
import requests

wrk_idades = []
wrk_refeicoes = []
wrk_cardapios = {}

xls_file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'tmp'))

# ..Meses e dias da semana
meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junio', 'Julio', 'Agosto', 'Setembro', 'Outubro',
         'Novembro', 'Dezemnbro']
dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']


def get_num_semana(s_data):
    dta = s_data[:4] + '-' + s_data[4:][:2] + '-' + s_data[6:]
    tz_dta = datetime.strptime(dta, '%Y-%m-%d')
    return (tz_dta.isocalendar()[1])


def estilo(linha, coluna, ws):
    bd = Side(style='thin', color="000000")
    ws.cell(row=linha, column=coluna).border = Border(left=bd, top=bd, right=bd, bottom=bd)
    ws.cell(row=linha, column=coluna).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def format_cels(refeicoes_idade, ws):
    n_refeicoes = sum(refeicoes_idade.values())

    # ..Bordas da da planilha
    for lin in range(1, 9):
        for col in range(1, n_refeicoes + 2):
            estilo(lin, col, ws)

    # ..Merge celulas do título Dia
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    estilo(1, 1, ws)

    # ..Merge celulas do titulo da planilha
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=n_refeicoes + 1)
    estilo(1, 2, ws)

    # ..Merge celulas das faixas etárias
    coluna = 2
    for k in refeicoes_idade.keys():
        f_refeicoes = refeicoes_idade[k]
        ws.merge_cells(start_row=2, start_column=coluna, end_row=2, end_column=coluna + f_refeicoes - 1)
        estilo(2, coluna, ws)
        coluna += f_refeicoes


def titulos(idade_refeicoes, ws, titulo):
    ws.cell(row=1, column=1).value = 'DIA'
    ws.cell(row=1, column=2).value = titulo

    coluna = 2
    for k in idade_refeicoes.keys():
        f_refeicoes = len(idade_refeicoes[k])
        ws.cell(row=2, column=coluna).value = k
        nome_refeicoes(idade_refeicoes[k], coluna, ws)
        coluna += f_refeicoes


def nome_refeicoes(lista, coluna, ws):
    col_ini = 0
    for c in range(col_ini, len(lista)):
        ws.cell(row=3, column=coluna + c).value = lista[c]


def get_cardapios(unidade, data_de, data_ate, c_status):
    # ..Acesso ao banco de dados
    # client = pymongo.MongoClient('localhost', 27017)
    # db = client.pratoaberto
    # collection = db.cardapios

    client = pymongo.MongoClient(os.environ.get('MONGO_HOST'))
    db = client.pratoaberto
    collection = db.cardapios

    cursor = collection.find({"tipo_unidade": unidade,
                              "data": {"$gte": data_de, "$lte": data_ate},
                              "status": c_status}).sort([("data", 1)])
    return cursor


def get_cardapio_unidade_especial_by_api(unidade, inicio, fim):
    api = os.environ.get('PRATOABERTO_API')
    # api = 'http://localhost:8000'
    url = api + '/editor/cardapios-unidade-especial/?unidade={}&inicio={}&fim={}'.format(unidade, inicio, fim)
    r = requests.get(url)

    return r.json()


def get_unidade_especial_by_id(id_ue):
    api = os.environ.get('PRATOABERTO_API')
    # api = 'http://localhost:8000'
    url = api + '/editor/unidade-especial/{}'.format(id_ue)
    r = requests.get(url)

    return r.json()


def set_cardapio(cardapios, doc_cardapio, doc_idade):
    l_refeicoes = list(doc_cardapio.keys())
    for refeicao in l_refeicoes:
        alimentos = ', '.join(doc_cardapio[refeicao])
        if doc_idade in cardapios:
            if refeicao in cardapios[doc_idade]:
                cardapios[doc_idade][refeicao] = alimentos


def save_dia(wb, ws, lin, data_ant, cardapios, xls_file):
    numero_dia = datetime.weekday(datetime(int(data_ant[:4]), int(data_ant[4:][:2]), int(data_ant[6:])))
    nome_dia = dias_semana[numero_dia]
    dia = nome_dia + ' ' + data_ant[6:] + '/' + meses[int(data_ant[4:][:2]) - 1][:3]
    ws.cell(row=lin, column=1).value = dia

    col = 2
    for idade in cardapios.keys():
        for refeicao in cardapios[idade]:
            if len(cardapios[idade][refeicao]) > 0:
                ws.cell(row=lin, column=col).value = cardapios[idade][refeicao]
            else:
                ws.cell(row=lin, column=col).value = "N/D"
            col += 1

    save_xls(wb, xls_file)


def set_planilha(unidade, menu, semana, data_de, data_ate, wb, ws, xls_file):
    semana_de, semana_ate = str(get_num_semana(data_de)), str(get_num_semana(data_ate))
    lista_semanas = [i for i in range(int(semana_de), int(semana_ate) + 1)]
    w_de, w_ate = str(get_num_semana(data_de)), str(get_num_semana(data_ate))

    dta_de = data_de[6:] + '-' + data_de[4:][:2] + '-' + data_de[:4]
    dta_ate = data_ate[6:] + '-' + data_ate[4:][:2] + '-' + data_ate[:4]

    semanas = ' - Semana: ' + w_de
    if len(lista_semanas) > 1:
        semanas = ' - Semanas: ' + w_de + ' a ' + w_ate

    titulo = 'CARDÁPIOS ' + unidade + ' - Período: ' + dta_de + ' até ' + dta_ate + semanas

    # num_recs = 0
    # cursor.rewind()
    # doc = cursor.next()
    # cursor.rewind()

    # while num_recs <= cursor.count():
    for doc in menu:
        if semana == get_num_semana(doc['data']):
            l_refeicoes = list(doc['cardapio'].keys())

            for refeicao in l_refeicoes:
                if doc['idade'] not in wrk_idades:
                    wrk_idades.append(doc['idade'])

                if refeicao not in wrk_refeicoes:
                    wrk_refeicoes.append(refeicao)

                if doc['idade'] not in wrk_cardapios:
                    wrk_cardapios[doc['idade']] = {}

                if refeicao not in wrk_cardapios[doc['idade']]:
                    wrk_cardapios[doc['idade']][refeicao] = ''

        # num_recs += 1

    #     try:
    #         doc = cursor.next()
    #     except Exception:
    #         num_recs += 1
    #         cursor.close()
    # cursor.close()

    # ..Ordena as faixas etárias e nome das refeições encontradas
    idades = sorted(wrk_idades)
    refeicoes = sorted(wrk_refeicoes)

    num_refeicoes_por_idade = {}
    cardapios = {}
    for k in idades:
        cardapios[k] = wrk_cardapios[k]
        num_refeicoes_por_idade[k] = len(cardapios[k])

    # ..Monta a lista das refeições por idade
    idade_refeicoes = {}
    lista = []
    for k in idades:
        for r in refeicoes:
            if r in cardapios[k]:
                lista.append(r)
        idade_refeicoes[k] = lista
        lista = []

    # ..Refaz cardapios: ordena o nome das refeicoes
    x_cardapios = {}
    for idade in idade_refeicoes.keys():
        if idade not in x_cardapios:
            x_cardapios[idade] = {}
        for r in idade_refeicoes[idade]:
            if r not in x_cardapios[idade]:
                x_cardapios[idade][r] = ''

    # ..Formata a planilha e preenche os tiltulos
    if len(num_refeicoes_por_idade) > 0:
        format_cels(num_refeicoes_por_idade, ws)
        titulos(idade_refeicoes, ws, titulo)

    save_xls(wb, xls_file)
    return x_cardapios


def clean_up(wb, xls_file):
    # ..Apaga planilhas sem cardápios
    for ws in wb.worksheets:
        semana = ws.title
        if ws['A4'].value is None:
            wb.remove(wb[semana])
            save_xls(wb, xls_file)


def save_xls(wb, xls_file):
    wb.save(xls_file)


def gera_excel(id_unidade):
    xls_file = None
    # ue_dict = ue.get_unidade(id_unidade)
    ue_dict = get_unidade_especial_by_id(id_unidade)
    unidade = ue_dict['nome']
    data_de = ue_dict['data_inicio']
    data_ate = ue_dict['data_fim']

    semana_de = str(get_num_semana(data_de))
    semana_ate = str(get_num_semana(data_ate))
    lista_semanas = [i for i in range(int(semana_de), int(semana_ate) + 1)]

    # cursor = get_cardapios(unidade, data_de, data_ate, 'PUBLICADO')
    menu = get_cardapio_unidade_especial_by_api(id_unidade, data_de, data_ate)

    if len(menu) == 0:
        print('Não há cardápios publicados no período {0} a {1} para a unidade {0}.'.format(data_de, data_ate, unidade))
        return False

    # ..Arquivo xls de saída
    # xls_file = xls_file_path / 'Cardapios_' + unidade + '_' + data_de + '_' + data_ate + '.xlsx'
    xls_file = '{}/{}_{}_{}_{}.xlsx'.format(xls_file_path, 'Cardapios_', unidade, data_de, data_ate)

    wb = Workbook()
    cardapios = {}

    # ..Cria as planilhas preparadas para receber os cardapios por dia.
    for semana in lista_semanas:
        ws = wb.create_sheet('Semana ' + str(semana))
        # cardapios = set_planilha(unidade, cursor, semana, data_de, data_ate, wb, ws, xls_file)
        cardapios = set_planilha(unidade, menu, semana, data_de, data_ate, wb, ws, xls_file)

    num_recs = 0
    lin = 4

    # cursor.rewind()
    # doc = cursor.next()
    # data_ant = doc['data']
    data_ant = menu[0]['data']
    # semana_ant = get_num_semana(doc['data'])
    semana_ant = get_num_semana(data_ant)
    ws = wb['Semana ' + str(semana_ant)]
    # cursor.rewind()

    # while num_recs < cursor.count():
    for doc in menu:
        # while data_ant == doc['data']:
        #     num_recs += 1
        set_cardapio(cardapios, doc[r'cardapio'], doc[r'idade'])
        # try:
        #     doc = cursor.next()
        # except Exception:
        #     save_dia(wb, ws, lin, data_ant, cardapios, xls_file)
        #     clean_up(wb, xls_file)
        #     return -1

        save_dia(wb, ws, lin, data_ant, cardapios, xls_file)

        if semana_ant != get_num_semana(doc['data']):
            semana_ant = get_num_semana(doc['data'])
            ws = wb['Semana ' + str(semana_ant)]
            lin = 3

        lin += 1
        data_ant = doc['data']

        # ..Limpa o conteúdo da estrutura
        l_idades = list(cardapios.keys())
        for idade in l_idades:
            l_refeicoes = list(cardapios[idade].keys())
            for refeicao in l_refeicoes:
                cardapios[idade][refeicao] = []

    set_cardapio(cardapios, doc['cardapio'], doc['idade'])
    save_dia(wb, ws, lin, data_ant, cardapios, xls_file)
    # cursor.close()
    clean_up(wb, xls_file)
    return xls_file


# - Main ---------------------------------------------------------------------------------------------------------------
if __name__ == "__main__":
    gera_excel("5cd412b6c9a1820007e2dd36")
    # gera_excel("5cb5ed424619f0726f26a353")
