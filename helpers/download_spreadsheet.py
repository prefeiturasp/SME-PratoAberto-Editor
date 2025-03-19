#

# download_spreadsheet.py
#
#   Gera arquivo excel com os cardápios publicados 
#   Ref: SME - Alimentação\Sprint 4 - Correções e Melhorias do Prato, item 663.
#
#   Uso: extrac_02 <periodo> <gestão> <tipo_escola> onde:
#
#                  <periodo>         "AAAAMMDD, AAAAMMDD"  - string com a data_de e data_até 
#                  <gestão>          - string com o tipo de gestão
#                  <tipo_escola>     - string com o tipo da escola
#
#  Nome do arquivo gerado: extracao_AAAAMMDDhhmm.xlsx'
#       onde: AAAA= ano  MM= mês  DD= dia  hh= hora  mm= minutos
#

import os
from copy import copy
from datetime import datetime
from pathlib import Path

import pymongo
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Border, Side

# ----------------------------------------------------------------------------------------------------------------------
#
#  Mapeamento das refeições por faixa etária vs unidade e gestão
# -

idades = {1: ['A - 0 A 1 MÊS', 'B - 1 A 3 MESES', 'C - 4 A 5 MESES', 'D - 6 A 7 MESES', 'E - 8 A 11 MESES',
              'X - 1A -1A E 11MES', 'F - 2 A 3 ANOS', 'G - 4 A 6 ANOS'],

          2: ['A - 0 A 1 MÊS', 'B - 1 A 3 MESES', 'C - 4 A 5 MESES', 'D - 6 A 7 MESES', 'E - 8 A 11 MESES',
              'X - 1A -1A E 11MES', 'F - 2 A 3 ANOS', 'G - 4 A 6 ANOS', 'H - ADULTO'],

          3: ['A - 0 A 1 MÊS', 'B - 1 A 3 MESES', 'C - 4 A 5 MESES', 'D - 6 A 7 MESES', 'E - 8 A 11 MESES',
              'X - 1A -1A E 11MES', 'F - 2 A 3 ANOS', 'G - 4 A 6 ANOS', 'W - EMEI DA CEMEI'],

          4: ['H - ADULTO', 'Z - UNIDADES SEM FAIXA'],

          5: ['D - 0 A 5 MESES', 'D - 6 MESES', 'D - 7 MESES', 'E - 8 A 11 MESES', 'X - 1A -1A E 11MESES',
              'I - 2 A 6 ANOS'],

          6: ['D - 0 A 5 MESES', 'D - 6 MESES', 'D - 7 MESES', 'E - 8 A 11 MESES', 'X - 1A -1A E 11MESES',
              'I - 2 A 6 ANOS', 'W - EMEI DA CEMEI'],

          7: ['Z - UNIDADES SEM FAIXA']}

refeicoes = {1: 'D - DESJEJUM',
             2: 'C - COLACAO',
             3: 'A - ALMOCO',
             4: 'L - LANCHE',
             5: 'J - JANTAR',
             6: 'AA - ALMOCO ADULTO',
             7: 'L4 - LANCHE 4 HORAS',
             8: 'L5 - LANCHE 5 HORAS',
             9: 'R1 - REFEICAO 1',
             10: 'R2 - REFEICAO 2',
             11: 'MS - MERENDA SECA',
             12: 'MI - MERENDA INICIAL',
             13: 'HE - HORARIO ESTENDIDO',
             14: 'LP - LANCHE PASSEIO',
             15: 'L - LANCHE CEMEI',
             16: 'R - REFEICAO CEMEI',
             17: 'L4 - ESCOLA ABERTA',
             18: 'L5 - ESCOLA ABERTA',
             19: 'R - ESCOLA ABERTA',
             20: 'R - REFEICAO',
             'WD': ['D - DESJEJUM',
                    'A - ALMOCO'],
             'WT': ['L5 - LANCHE 5 HORAS',
                    'R1 - REFEICAO 1']}

unidades = {'CEI_MUNICIPAL': {
    'DIRETA': {'fx_etaria': 1, 'refeicao': [1, 2, 3, 4, 5], 'template': 'Template_1.xlsx'},
    'MISTA': {'fx_etaria': 1, 'refeicao': [1, 2, 3, 4, 5], 'template': 'Template_1.xlsx'},
    'TERCEIRIZADA': {'fx_etaria': 5, 'refeicao': [1, 2, 3, 4, 5], 'template': 'Template_5.xlsx'}},

    'CCI': {
        'DIRETA': {'fx_etaria': 1, 'refeicao': [1, 2, 3, 4, 5], 'template': 'Template_1.xlsx'},
        'MISTA': {'fx_etaria': 1, 'refeicao': [1, 2, 3, 4, 5], 'template': 'Template_1.xlsx'},
        'TERCEIRIZADA': {'fx_etaria': 5, 'refeicao': [1, 2, 3, 4, 5], 'template': 'Template_5.xlsx'}},

    'CEI_PARCEIRO_(RP)': {
        'DIRETA': {'fx_etaria': 2, 'refeicao': [1, 2, 3, 4, 6], 'template': 'Template_2.xlsx'},
        'MISTA': {'fx_etaria': 1, 'refeicao': [1, 2, 3, 4, 6], 'template': 'Template_2.xlsx'}},

    'PROJETO_CECI': {
        'DIRETA': {'fx_etaria': 2, 'refeicao': [1, 2, 3, 4, 6], 'template': 'Template_2.xlsx'},
        'MISTA': {'fx_etaria': 1, 'refeicao': [1, 2, 3, 4, 6], 'template': 'Template_2.xlsx'}},

    'CEMEI': {
        'DIRETA': {'fx_etaria': 3, 'refeicao': [1, 2, 3, 4, 5, 'WD'], 'template': 'Template_3.xlsx'},
        'MISTA': {'fx_etaria': 3, 'refeicao': [1, 2, 3, 4, 5, 'WD'], 'template': 'Template_3.xlsx'},
        'TERCEIRIZADA': {'fx_etaria': 6, 'refeicao': [1, 2, 3, 4, 5, 'WT'], 'template': 'Template_6.xlsx'}},

    'EMEI': {
        'DIREITA': {'fx_etaria': 4, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_4.xlsx'},
        'MISTA': {'fx_etaria': 4, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_4.xlsx'},
        'TERCEIRIZADA': {'fx_etaria': 7, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_7.xlsx'}},

    'EMEF': {
        'DIRETA': {'fx_etaria': 4, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_4.xlsx'},
        'MISTA': {'fx_etaria': 4, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_4.xlsx'},
        'TERCEIRIZADA': {'fx_etaria': 7, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_7.xlsx'}},

    'CIEJA': {
        'DIRETA': {'fx_etaria': 4, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_4.xlsx'},
        'MISTA': {'fx_etaria': 4, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_4.xlsx'},
        'TERCEIRIZADA': {'fx_etaria': 7, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_7.xlsx'}},

    'EMEBS': {
        'DIRETA': {'fx_etaria': 4, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_4.xlsx'},
        'MISTA': {'fx_etaria': 4, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_4.xlsx'},
        'TERCEIRIZADA': {'fx_etaria': 7, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_7.xlsx'}},

    'SME_CONVÊNIO': {
        'DIRETA': {'fx_etaria': 4, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_4.xlsx'},
        'MISTA': {'fx_etaria': 4, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_4.xlsx'},
        'TERCEIRIZADA': {'fx_etaria': 7, 'refeicao': [11, 6, 7, 8, 9, 10], 'template': 'Template_7.xlsx'}}
}

# ----------------------------------------------------------------------------------------------------------------------
# ..Meses e dias da semana
meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junio', 'Julio', 'Agosto', 'Setembro', 'Outubro',
         'Novembro', 'Dezemnbro']
dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']


# ----------------------------------------------------------------------------------------------------------------------
def get_num_semana(s_data):
    dta = s_data[:4] + '-' + s_data[4:][:2] + '-' + s_data[6:]
    tz_dta = datetime.strptime(dta, '%Y-%m-%d')
    return (tz_dta.isocalendar()[1])


# ----------------------------------------------------------------------------------------------------------------------
def estilo(n_cels, pl):
    bd = Side(style='thick', color="000000")
    c = 2
    while c < n_cels:
        pl.cell(row=1, column=c).border = Border(left=bd, top=bd, right=bd, bottom=bd)
        pl.cell(row=3, column=c).border = Border(left=bd, top=bd, right=bd, bottom=bd)
        c += 1


# ----------------------------------------------------------------------------------------------------------------------
def merge_cels(n_faixas, n_refeicoes, pl):
    # ..Dia
    pl.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)

    # ..Titulo
    n_col = (n_faixas * n_refeicoes) + 1
    pl.merge_cells(start_row=1, start_column=2, end_row=1, end_column=n_col)

    # ..Faixas etárias
    fx = 0
    while fx < n_faixas:
        pl.merge_cells(start_row=2, start_column=2 + (n_refeicoes * fx), end_row=2,
                       end_column=1 + n_refeicoes + (fx * n_refeicoes))
        fx += 1


# ----------------------------------------------------------------------------------------------------------------------
def titulos(n_faixas, n_refeicoes, escola, gestao, dt_publicacao, pl):
    f = 0
    fx = 0
    while f < (n_faixas * n_refeicoes):
        pl.cell(row=2, column=2 + f).value = idades[unidades[escola][gestao]['fx_etaria']][
                                                 fx] + ' Publicação: ' + dt_publicacao
        c = 0
        while c < n_refeicoes:
            pl.cell(row=3, column=c + 2 + f).value = refeicoes[unidades[escola][gestao]['refeicao'][c]]
            c += 1
        f += n_refeicoes
        fx += 1


# ----------------------------------------------------------------------------------------------------------------------
def save_dia(sv_cardapios_dia, sv_lin, sv_data_ant, sv_num_faixas, sv_num_refeicoes, sv_tipo_escola, sv_tipo_gestao,
             sv_dta_pub, sv_wb, sv_ws, sv_template, sv_cabecalho):
    # ..Titulos e formatação
    sv_ws.cell(row=1, column=2).value = sv_cabecalho
    titulos(sv_num_faixas, sv_num_refeicoes, sv_tipo_escola, sv_tipo_gestao, sv_dta_pub, sv_ws)
    estilo((sv_num_refeicoes * sv_num_faixas) + 2, sv_ws)
    sv_wb.save(sv_template)

    # ..Monta o conteúdo da celula Dia e popula a celula
    num_dia = datetime.weekday(datetime(int(sv_data_ant[:4]), int(sv_data_ant[4:][:2]), int(sv_data_ant[6:])))
    nom_dia = dias_semana[num_dia]
    dia = nom_dia + ' ' + sv_data_ant[6:] + '/' + meses[int(sv_data_ant[4:][:2]) - 1][:3]
    sv_ws.cell(row=sv_lin, column=1).value = dia

    # ..Popula as celulas da planilha com os cardápios
    col = 2
    n_faixa = 0
    n_refeicao = 0
    while n_faixa < sv_num_faixas:
        while n_refeicao < sv_num_refeicoes:
            ref_n = refeicoes[unidades[sv_tipo_escola][sv_tipo_gestao][r'refeicao'][n_refeicao]]
            if len(sv_cardapios_dia[idades[unidades[sv_tipo_escola][sv_tipo_gestao][r'fx_etaria']][n_faixa]][
                       ref_n]) > 0:
                sv_ws.cell(row=sv_lin, column=col).value = \
                    sv_cardapios_dia[idades[unidades[sv_tipo_escola][sv_tipo_gestao][r'fx_etaria']][n_faixa]][ref_n][0]
            else:
                sv_ws.cell(row=sv_lin, column=col).value = 'N/D'
            col += 1
            n_refeicao += 1
        n_faixa += 1
        n_refeicao = 0


# ----------------------------------------------------------------------------------------------------------------------
def gera_excel(parametros):
    try:
        # ..Acesso ao banco de dados
        client = pymongo.MongoClient(os.environ.get('MONGO_HOST'))
        db = client.pratoaberto
        collection = db.cardapios

        # ..Data e hora da emissão
        data = (datetime.now().isoformat(timespec='minutes')).split('T')
        hoje, hora = data[0], data[1]
        dt = ''.join(x for x in hoje.split('-'))
        hr = ''.join(x for x in hora.split(':'))

        # ..Parámetros para a extração e gravação
        data_de = parametros.split(',')[0].strip()
        data_ate = parametros.split(',')[1].strip()
        tipo_gestao = parametros.split(',')[2].strip()
        tipo_escola = parametros.split(',')[3].strip()
        t_status = 'PUBLICADO'

        # ..Arquivo Template
        arq_nome = unidades[tipo_escola][tipo_gestao]['template']
        template = Path(os.path.abspath('static/spreadsheet')) / arq_nome

        # ..Arquivo xlsx de saída
        arq_nome = 'Cardapios_' + tipo_escola + '_' + dt + hr + '.xlsx'
        xls = Path(os.path.abspath('arquivos/')) / arq_nome

        # ..Datas do período e número de semanas
        dta_de = data_de[6:] + '-' + data_de[4:][:2] + '-' + data_de[:4]
        dta_ate = data_ate[6:] + '-' + data_ate[4:][:2] + '-' + data_ate[:4]
        w_de, w_ate = str(get_num_semana(data_de)), str(get_num_semana(data_ate))

        # ..Elementos da planilha
        num_faixas = len(idades[unidades[tipo_escola][tipo_gestao]['fx_etaria']])
        num_refeicoes = len(unidades[tipo_escola][tipo_gestao]['refeicao'])
        cabecalho = 'CARDÁPIOS ' + tipo_escola + ' - ' + tipo_gestao + ' - Período: ' + dta_de + ' até ' + dta_ate + ' - Semanas: ' + w_de + ' a ' + w_ate

        # ..Verifica a existência de cardápios com os parámetros fornecidos
        consulta = [{"$match": {"tipo_unidade": tipo_escola, "data": {"$gte": data_de, "$lte": data_ate},
                                "status": t_status, "tipo_atendimento": tipo_gestao,
                                "data_publicacao": {"$exists": "true"}}}, {"$sort": {"data": 1}},
                    {"$count": "num_regs"}]

        num_regs = 0
        for i in collection.aggregate(consulta):
            num_regs = i['num_regs']

        num_semanas = int(w_ate) - int(w_de) + 1
        semana = 0
        aba = 0
        new_wb = None

        # ..Cria as abas das semanas
        if num_regs > 0:
            while semana < num_semanas:
                try:
                    wb = load_workbook(filename=template)
                    if num_semanas > 1:
                        if aba + 2 > num_semanas:
                            aba = num_semanas
                        else:
                            new_wb = wb.create_sheet('Semana ' + str(aba + 2))
                            aba += 1

                            default_sheet = wb['Semana 1']

                            for row in default_sheet.rows:
                                for cell in row:
                                    if isinstance(cell, Cell):
                                        new_cell = new_wb.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                                        if cell.has_style:
                                            new_cell.font = copy(cell.font)
                                            new_cell.border = copy(cell.border)
                                            new_cell.fill = copy(cell.fill)
                                            new_cell.number_format = copy(cell.number_format)
                                            new_cell.protection = copy(cell.protection)
                                            new_cell.alignment = copy(cell.alignment)
                    else:
                        default_sheet = wb['Semana 1']

                except Exception as e:
                    print(e)
                    print('Erro no bloco de criação do arquivo template'.format(template))
                    return -1

                # ..Formata celulas dos titulos
                if num_semanas > 1:
                    merge_cels(num_faixas, num_refeicoes, new_wb)

                # ..Grava aba da semana no teplate
                arq_nome = 'x' + unidades[tipo_escola][tipo_gestao]['template']
                template = Path(os.path.abspath('static/spreadsheet')) / arq_nome

                wb.save(template)
                semana += 1

            # ..Abre o template com todas as semanas
            wb = load_workbook(filename=template)

            # ..Formata aba 'Semana 1'
            ws = wb['Semana 1']
            merge_cels(num_faixas, num_refeicoes, ws)
            wb.save(template)
            wb = load_workbook(filename=template)

            # ..Inicia a estrutura auxiliar para receber os cardápios do dia
            cardapios_dia = {}
            n_faixa = 0
            n_refeicao = 0
            while n_faixa < num_faixas:
                cardapios_dia[idades[unidades[tipo_escola][tipo_gestao]['fx_etaria']][n_faixa]] = {}
                while n_refeicao < num_refeicoes:
                    ref_n = refeicoes[unidades[tipo_escola][tipo_gestao]['refeicao'][n_refeicao]]
                    cardapios_dia[idades[unidades[tipo_escola][tipo_gestao]['fx_etaria']][n_faixa]][ref_n] = []
                    n_refeicao += 1
                n_faixa += 1
                n_refeicao = 0

            # ..Extrai os cardápios do banco de dados
            cursor = collection.find({"tipo_atendimento": tipo_gestao, "tipo_unidade": tipo_escola, "status": t_status,
                                      "data_publicacao": {"$exists": "true"},
                                      "data": {"$gte": data_de, "$lte": data_ate}}).sort([("data", 1)])

            # ..Processa os cardápios extraidos
            num_recs = 0
            doc = cursor.next()
            data_ant = doc['data']
            semana_ant = get_num_semana(data_ant)
            num_planilha = 1
            lin = 4

            cursor.rewind()
            while num_recs <= cursor.count():
                if semana_ant == get_num_semana(doc['data']):
                    if data_ant == doc['data']:
                        l_refeicoes = list(doc[r'cardapio'].keys())
                        for refeicao in l_refeicoes:
                            alimentos = ', '.join(doc[r'cardapio'][refeicao])
                            if doc[r'idade'] in cardapios_dia:
                                if refeicao in cardapios_dia[doc[r'idade']]:
                                    cardapios_dia[doc[r'idade']][refeicao].append(alimentos)

                        num_recs += 1

                        try:
                            doc = cursor.next()
                        except Exception:
                            num_recs += 1
                            cursor.close()

                    else:
                        # ..Popula planilha com o cardápio do dia
                        ws = wb['Semana ' + str(num_planilha)]
                        dta_pub = doc['data_publicacao'][:10]

                        # ..Grava o dia na planilha
                        save_dia(cardapios_dia, lin, data_ant, num_faixas, num_refeicoes, tipo_escola, tipo_gestao,
                                 dta_pub, wb, ws, template, cabecalho)
                        lin += 1
                        wb.save(template)

                        # ..Limpa o conteúdo da estrutura auxiliar
                        l_idades = list(cardapios_dia.keys())
                        for idade in l_idades:
                            l_refeicoes = list(cardapios_dia[idade].keys())
                            for refeicao in l_refeicoes:
                                cardapios_dia[idade][refeicao] = []

                        data_ant = doc['data']

                else:
                    # ..Grava o último dia da semana
                    ws = wb['Semana ' + str(num_planilha)]
                    dta_pub = doc['data_publicacao'][:10]

                    save_dia(cardapios_dia, lin, data_ant, num_faixas, num_refeicoes, tipo_escola, tipo_gestao, dta_pub,
                             wb, ws, template, cabecalho)

                    wb.save(template)

                    # ..Continua o processo com a seguinte semana
                    semana_ant = get_num_semana(doc['data'])
                    data_ant = doc['data']
                    num_planilha += 1
                    lin = 4
            # ..end while

            # ..Salva o último dia da semana
            ws = wb['Semana ' + str(num_planilha)]
            dta_pub = doc['data_publicacao'][:10]
            save_dia(cardapios_dia, lin, data_ant, num_faixas, num_refeicoes, tipo_escola, tipo_gestao, dta_pub, wb, ws,
                     template, cabecalho)
            wb.save(template)

            # ..Apaga planilhas sem cardápios
            num_semanas = int(w_ate) - int(w_de) + 1
            semana = 1

            while semana <= num_semanas:
                ws = wb['Semana ' + str(semana)]
                v = ws['B3'].value
                if v == None:
                    wb.remove(wb['Semana ' + str(semana)])
                semana += 1

            wb.save(xls)
            return xls
        else:
            print('Não há cardápios cadastrados nesse período.')

    except Exception as e:
        print(e)
        print('Não foi possível gerar o arquivo excel.')

# - Main ---------------------------------------------------------------------------------------------------------------
# if __name__ == "__main__":
#
#     if len(sys.argv) != 4:
#         print(len(sys.argv))
#         print('Uso: extrac_cardapios <periodo> <gestão> <tipo_escola>   onde:')
#         print('          periodo = string: "AAAAMMDD, AAAAMMDD"')
#         print('          gestão  = string: tipo de gestão')
#         print('      tipo_escola = string: tipo da escola')
#     else:
#         params = [p for p in sys.argv]
#         params.__delitem__(0)
#
#     gera_excel(params)
