import json
import os
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment

XLSX_FILE_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'tmp'))
DATA_GERACAO = str(datetime.now()).replace(' ', '_').replace(':', '_').replace('.', '_')
bd = Side(style='thin', color="000000")
BORDER = Border(left=bd, top=bd, right=bd, bottom=bd)
ALIGNMENT = Alignment(horizontal='center', vertical='center')
ARQUIVO = os.path.abspath(os.path.join(os.path.dirname(__file__), 'ordenacao.json'))
with open(ARQUIVO, 'r') as f:
    JSON = json.load(f)


class GeradorExcelCardapioUE(object):

    def __init__(self, id_unidade_especial):
        self.api = os.environ.get('PRATOABERTO_API')
        # self.api = 'http://localhost:8000'
        self.id_unidade_especial = id_unidade_especial
        self.data_inicio = None
        self.data_final = None
        self.cardapios = []

    def _get_unidade_especial(self):
        url = '{}/editor/unidade-especial/{}'.format(self.api, self.id_unidade_especial)
        response = requests.get(url)
        if response.status_code == 200:
            return response.json()
        return {}

    def _get_cardapio_unidade_especial(self):
        url = '{}/editor/cardapios-unidade-especial/?unidade={}&inicio={}&fim={}'.format(self.api,
                                                                                         self.id_unidade_especial,
                                                                                         self.data_inicio,
                                                                                         self.data_final)
        response = requests.get(url)
        if response.status_code == 200:
            return self._extrai_dados_importantes(response.json())
        return {}

    def _extrai_dados_importantes(self, cardapios):
        novo_cardapio = []
        if cardapios:
            for cardapio in cardapios:
                novos_dados = {'semana': self._get_numero_semana_ano(cardapio.get('data')),
                               'faixa': cardapio.get('idade'),
                               'data': cardapio.get('data'),
                               'polo': cardapio.get('tipo_unidade'),
                               'cardapio': cardapio.get('cardapio')}
                novo_cardapio.append(novos_dados)
        return sorted(novo_cardapio, key=lambda v: v.get('data'))

    def _get_numero_semana_ano(self, str_date):
        try:
            semana = datetime.strptime(str_date, '%Y%m%d').isocalendar()[1]
            return semana
        except ValueError:
            return False

    def _get_abas_planinha(self):
        semanas = []
        for cardapio in self.cardapios:
            semana_ano = self._get_numero_semana_ano(cardapio.get('data'))
            if semana_ano not in semanas:
                semanas.append(semana_ano)
        return sorted(semanas)

    def _get_faixas_etarias_by_semana(self, semana):
        faixas_set = set()
        [faixas_set.add(cardapio.get('faixa')) for cardapio in self.cardapios if cardapio.get('semana') == semana]
        return list(faixas_set)

    def _get_datas_cardapios(self, semana):
        datas_lista = []
        [datas_lista.append(cardapio.get('data')) for cardapio in self.cardapios if
         cardapio.get('semana') == semana and (cardapio.get('data') not in datas_lista)]
        return datas_lista

    def _criar_estrutura_planilha(self):
        workbook = Workbook()
        semanas = self._get_abas_planinha()
        titulo_polo = self.cardapios[0]['polo']
        contador_coluna = 0
        for semana in semanas:
            aba_planilha = workbook.create_sheet('Semana {}'.format(semana), contador_coluna)
            contador_coluna += 1
            faixa_etarias = self._get_faixas_etarias_by_semana(semana)
            categorias_faixas_dict = self._get_categorias_by_faixa_etaria_dict(semana, faixa_etarias)
            datas_cardapio = self._get_datas_cardapios(semana)
            self._criar_titulos(aba_planilha, semanas, titulo_polo)
            self._criar_faixa_etaria_com_categoria(categorias_faixas_dict, aba_planilha)
            self._criar_datas_cardapio(datas_cardapio, aba_planilha)
            self._criar_refeicoes_cardapios(semana, categorias_faixas_dict, datas_cardapio, aba_planilha)
        arquivo_excel = '{}/CARDAPIOS_UNIDADE_ESPCIAL_{}.xlsx'.format(XLSX_FILE_PATH, DATA_GERACAO)
        workbook.save(arquivo_excel)
        return arquivo_excel

    def _criar_refeicoes_cardapios(self, semana, categorias_faixas_dict, datas_cardapio, aba_planilha):
        posicao_linha_inicial = 4
        posicisao_coluna_inicial = 2
        contador_colunas_faixa = 0
        for faixa, categorias in categorias_faixas_dict.items():
            quantidade_categoria = len(categorias)
            for data_cardapio in datas_cardapio:
                for categoria in categorias:
                    refeicao = self.get_refeicao_by_categoria(faixa, data_cardapio, categoria, semana)
                    self.adicionando_refeicao_planilha(refeicao, posicao_linha_inicial, posicisao_coluna_inicial,
                                                       aba_planilha)
                    posicisao_coluna_inicial += 1
                posicao_linha_inicial += 1
                if contador_colunas_faixa > quantidade_categoria:
                    posicisao_coluna_inicial = quantidade_categoria + 2
                else:
                    posicisao_coluna_inicial = 2
                contador_colunas_faixa += 1
            posicisao_coluna_inicial = quantidade_categoria + 2
            posicao_linha_inicial = 4

    def adicionando_refeicao_planilha(self, refeicao, linha, coluna, aba_planilha):
        celula_refeicao = aba_planilha.cell(row=linha, column=coluna)
        celula_refeicao.value = ', '.join(refeicao)
        celula_refeicao.alignment = ALIGNMENT
        celula_refeicao.border = BORDER

    def get_refeicao_by_categoria(self, faixa, data, categoria, semana):
        for cardapio in self.cardapios:
            if cardapio.get('data') == data and cardapio.get('faixa') == faixa and cardapio.get('semana') == semana:
                for categ, refeicao in cardapio.get('cardapio').items():
                    if categ == categoria:
                        return refeicao
        return []

    def _criar_titulos(self, aba_planilha, semanas, titulo_polo):
        data_inicio_formatada = self._converte_str_to_date(self.data_inicio, format='%d/%m/%Y')
        data_fim_formatada = self._converte_str_to_date(self.data_final, format='%d/%m/%Y')
        titulo_semanas = ' à '.join([str(value) for value in semanas])
        self._escrever_titulos_planilha(aba_planilha, data_fim_formatada, data_inicio_formatada, titulo_polo,
                                        titulo_semanas)

    def _escrever_titulos_planilha(self, aba_planilha, data_fim_formatada, data_inicio_formatada, titulo_polo,
                                   titulo_semanas):
        aba_planilha.merge_cells('A1:A3')
        celula_dia = aba_planilha.cell(row=1, column=1)
        celula_dia.value = 'DIA'
        aba_planilha.merge_cells('B1:F1')
        celula_titulo = aba_planilha.cell(row=1, column=2)
        celula_titulo.value = 'CARDÁPIOS {} - Período: {} até {} - Semanas: {}'.format(titulo_polo,
                                                                                       data_inicio_formatada,
                                                                                       data_fim_formatada,
                                                                                       titulo_semanas)
        celula_dia.border = BORDER
        celula_dia.alignment = ALIGNMENT
        celula_titulo.alignment = ALIGNMENT
        celula_titulo.border = BORDER

    def _criar_datas_cardapio(self, datas_cardapio, aba_planilha):
        contador_linha = 4
        for data in datas_cardapio:
            data_formatada = self._formatar_data_planilha(data)
            coluna_dias = aba_planilha.cell(row=contador_linha, column=1)
            coluna_dias.value = data_formatada
            coluna_dias.border = BORDER
            coluna_dias.alignment = Alignment(horizontal='center', vertical='center')
            contador_linha += 1

    def _ordena_dicionario_cardapio(self, cardapio_dict):
        lista_ordenada = []
        [lista_ordenada.append(val[key]) for key, value in JSON['refeicoes'].items() for val in cardapio_dict if
         key in val]
        return lista_ordenada

    def _formatar_data_planilha(self, data_planilha):
        data = datetime.strptime(data_planilha, '%Y%m%d')
        dia_semana = self._get_dia_semana(data.weekday())
        mes = self._get_mes(data.month)
        return '%s %02d/%s' % (dia_semana, data.day, mes)

    def _get_dia_semana(self, dia_semana):
        semana = {
            0: 'Segunda',
            1: 'Terça',
            2: 'Quarta',
            3: 'Quinta',
            4: 'Sexta',
            5: 'Sábado',
            6: 'Domingo',
        }
        return semana[dia_semana]

    def _get_mes(self, numero_mes):
        meses = {
            1: 'Jan',
            2: 'Fev',
            3: 'Mar',
            4: 'Abr',
            5: 'Mai',
            6: 'Jun',
            7: 'Jul',
            8: 'Ago',
            9: 'Set',
            10: 'Out',
            11: 'Nov',
            12: 'Dez'
        }
        return meses[numero_mes]

    def _converte_str_to_date(self, data_inicio, format='%Y%m%d'):
        data = datetime.strptime(data_inicio, '%Y%m%d')
        return datetime.strftime(data, format)

    def produzir_excel(self):
        unidade_especial = self._get_unidade_especial()
        self.data_inicio = unidade_especial['data_inicio']
        self.data_final = unidade_especial['data_fim']
        self.cardapios = self._get_cardapio_unidade_especial()
        if len(self.cardapios):
            return self._criar_estrutura_planilha()
        return False

    def _get_categorias_by_faixa_etaria_dict(self, semana, faixa_etarias):
        categoria_dict = {}
        faixas_etarias_ordenada = sorted(faixa_etarias, key=lambda k: k)
        for faixa in faixas_etarias_ordenada:
            categoria_dict[faixa] = []
            for cardapio in self.cardapios:
                if cardapio.get('semana') == semana and cardapio.get('faixa') == faixa:
                    categoria_dict[faixa] = cardapio.get('cardapio').keys()
        return categoria_dict

    def _criar_faixa_etaria_com_categoria(self, categorias_faixas_dict, aba_planilha):
        coluna_inicial_faixa = 2
        coluna_inicial_categoria = 2
        for key, values in categorias_faixas_dict.items():
            if coluna_inicial_faixa > 2:
                coluna_final = (coluna_inicial_faixa + len(values) - 1)
            else:
                coluna_final = len(values) + 1
            coluna_inicial_faixa = self._criar_faixa_etaria_categoria_planilha(aba_planilha, coluna_final,
                                                                               coluna_inicial_categoria,
                                                                               coluna_inicial_faixa, key, values)

    def _criar_faixa_etaria_categoria_planilha(self, aba_planilha, coluna_final, coluna_inicial_categoria,
                                               coluna_inicial_faixa, key, values):
        aba_planilha.merge_cells(start_row=2, end_row=2, start_column=coluna_inicial_faixa, end_column=coluna_final)
        linha_faixa = aba_planilha.cell(row=2, column=coluna_inicial_faixa)
        linha_faixa.value = str(key)
        linha_faixa.alignment = ALIGNMENT
        linha_faixa.border = BORDER
        for categoria in values:
            linha_categoria = aba_planilha.cell(row=3, column=coluna_inicial_categoria)
            linha_categoria.value = categoria
            linha_categoria.alignment = ALIGNMENT
            linha_categoria.border = BORDER
            coluna_inicial_categoria += 1
        coluna_inicial_faixa = coluna_inicial_categoria
        return coluna_inicial_faixa
